#!/usr/bin/env python3
"""
build_fmv.py — Robograder FMV index converter (canonical build tool, MERGE-SAFE).

Converts the data-entry workbook (Robograder_Index_Input_N.xlsx) into fmv_comics.json.
Unlike the old converter, this MERGES onto the existing fmv_comics.json instead of
rebuilding from scratch:
  * final books = the current fmv_comics.json books, OVERLAID with the workbook's
    rows (workbook wins for any key it defines). So books added through the admin
    dashboard (not in the workbook) are PRESERVED across a rebuild.
  * volumes (the volume/year model) is carried through untouched.
  * volumeGuards = current guards, unioned with the curated guards below and any
    Volume-column guards (curated/new win). So Batman etc. get added without
    dropping the Silver Surfer guards that only exist in the current file.
Cells are tier numbers 1-13. Key = fmvKey (mirrors index.html _fmvKeyOf).

USAGE:  python3 build_fmv.py [INPUT.xlsx] [CURRENT_fmv_comics.json] [OUT.json]
"""
import openpyxl, json, re, sys

IN_XLSX  = sys.argv[1] if len(sys.argv)>1 else 'Robograder_Index_Input_13.xlsx'
OLD_JSON = sys.argv[2] if len(sys.argv)>2 else 'fmv_comics.json'
OUT_JSON = sys.argv[3] if len(sys.argv)>3 else 'fmv.new.json'
OLD=json.load(open(OLD_JSON)); TIERS=OLD['tiers']

GUARD_CUTOFFS = { 'silver surfer': { '1968': 1971 } }
CURATED_GUARDS = {
  "amazing fantasy|15":1963, "avengers|1":1964, "daredevil|1":1965, "fantastic four|1":1962,
  "house of secrets|92":1972, "incredible hulk|181":1975, "moon knight|1":1981, "nova|1":1977,
  "predator|1":1990, "thundercats|1":1986, "uncanny x-men|1":1964, "wolverine|8":1990,
  "batman|1":1941, "batman|2":1941, "batman|3":1941, "batman|4":1941,
  "batman|5":1942, "batman|6":1942, "batman|7":1942, "batman|8":1942,
  "batman|9":1943, "batman|10":1943, "batman|121":1960, "batman|181":1967,
  "batman|189":1968, "batman|227":1971, "batman|251":1974, "batman|423":1989, "batman|457":1991,
}
# ASM vol-1 ran issues 1-441 (1963-1998). Modern volumes (2014/2018/2022) restarted at
# low numbers that COLLIDE with these vintage keys, so a post-1998 ASM #1-441 is a
# different book -> guard to the modern blanket. #601 (2009) and annuals (A#) are NOT
# vol-1 renumber collisions, so they are left unguarded.
for _asm in range(1, 442):
    CURATED_GUARDS[f"amazing spider-man|{_asm}"] = 1998

def normTitle(s):
    if not s: return ''
    return re.sub(r'^the\s+','',re.sub(r'\s+',' ',str(s).strip()),flags=re.I).lower()
def coerce_issue(issue):
    if issue is None: return ''
    if isinstance(issue,float) and issue.is_integer(): return str(int(issue))
    if isinstance(issue,int): return str(issue)
    return re.sub(r'^(\d+)\.0+$', r'\1', str(issue).strip())
def fmvKey(title,issue):
    t=re.sub(r'\s+',' ',str(title or '').strip()); i=re.sub(r'^#','',coerce_issue(issue))
    ta=re.search(r'\bannual\b',t,re.I)
    ma=re.match(r'^\s*(?:annual|ann\.?)\s*#?\s*0*(\d+)\s*$',i,re.I)
    mA=re.match(r'^\s*A\s*0*(\d+)\s*$',i,re.I)
    if ta:
        t=re.sub(r'\s+',' ',re.sub(r'\bannual\b','',t,flags=re.I)).strip()
        num=ma.group(1) if ma else (mA.group(1) if mA else re.sub(r'^0+(\d)',r'\1',i)); i='A'+num
    elif ma: i='A'+ma.group(1)
    elif mA: i='A'+mA.group(1)
    else: i=re.sub(r'^0+(\d)',r'\1',i)
    t=re.sub(r'^invincible\s+iron\s+man\b','Iron Man',t,flags=re.I)
    return normTitle(t)+'|'+i

# ---- current index -> base book map (key -> curve array) ----
def deref(v): return OLD['curves'][v] if isinstance(v,int) else v
base_books={k:[list(p) for p in deref(v)] for k,v in OLD['books'].items()}

# ---- workbook -> overlay books ----
wb=openpyxl.load_workbook(IN_XLSX, read_only=True, data_only=True)
def cell(r,i): return r[i] if i<len(r) else None
CATCHALLS=['Other Value Keys','EC Comics','Most Submitted','All Others']
SERIES=[n for n in wb.sheetnames if n not in (['README']+CATCHALLS)]
ORDER=SERIES+[c for c in CATCHALLS if c in wb.sheetnames]   # skip catch-alls that don't exist

wb_books={}; origin={}; dupes=[]; badtier=[]
for name in ORDER:
    if name not in wb.sheetnames: continue
    rows=list(wb[name].iter_rows(values_only=True))
    if not rows: continue
    hdr=rows[0]
    grade_cols=[(i,float(h)) for i,h in enumerate(hdr) if isinstance(h,(int,float)) or (isinstance(h,str) and re.fullmatch(r'\d+(\.\d+)?',(h or '').strip()))]
    title_i=next((i for i,h in enumerate(hdr) if isinstance(h,str) and (h or '').strip().lower()=='title'), None)
    issue_i=next((i for i,h in enumerate(hdr) if isinstance(h,str) and (h or '').strip().lower()=='issue'), None)
    if issue_i is None: continue
    for r in rows[1:]:
        iss=cell(r,issue_i)
        if iss in (None,''): continue
        title = cell(r,title_i) if title_i is not None else name
        if title in (None,''): continue
        curve=[]
        for i,g in grade_cols:
            v=cell(r,i)
            if v in (None,''): continue
            try: tv=int(round(float(v)))
            except: continue
            if tv<1 or tv>13: badtier.append((name,str(title),str(iss),g,v)); continue
            curve.append([g,tv])
        if not curve: continue
        curve.sort(key=lambda x:x[0])
        comp=[]; last=None
        for g,tv in curve:
            if tv!=last: comp.append([g,tv]); last=tv
        key=fmvKey(title,iss)
        if key in wb_books:
            dupes.append((key,name,'same' if wb_books[key]==comp else 'CONFLICT')); continue
        wb_books[key]=comp; origin[key]=(name,str(iss))

# ---- MERGE: current index as base, workbook overlaid on top ----
merged=dict(base_books); merged.update(wb_books)

# ---- guards: current guards, unioned with Volume-column + curated (curated win) ----
GUARDS=dict(OLD.get('volumeGuards',{}))
for name in SERIES:
    if name not in wb.sheetnames: continue
    rows=list(wb[name].iter_rows(values_only=True))
    if not rows: continue
    hdr=rows[0]
    voli=next((i for i,h in enumerate(hdr) if isinstance(h,str) and 'volume' in (h or '').lower()), None)
    issue_i=next((i for i,h in enumerate(hdr) if isinstance(h,str) and (h or '').strip().lower()=='issue'), None)
    if voli is None or issue_i is None: continue
    cut=GUARD_CUTOFFS.get(normTitle(name))
    if not cut: continue
    for r in rows[1:]:
        iss=cell(r,issue_i); vol=cell(r,voli)
        if iss in (None,'') or vol in (None,''): continue
        vl=str(vol).strip().replace('.0','')
        if vl in cut:
            k=fmvKey(name,iss)
            if k in merged: GUARDS[k]=cut[vl]
for _k,_y in CURATED_GUARDS.items():
    if _k in merged: GUARDS[_k]=_y

# ---- dedup curves ----
curve_index={}; curves=[]; booksOut={}
for k in sorted(merged):
    c=merged[k]; sig=json.dumps(c)
    if sig not in curve_index: curve_index[sig]=len(curves); curves.append(c)
    booksOut[k]=curve_index[sig]

# ---- version: bump patch of current ----
vparts=str(OLD.get('version','1.8.0')).split('.')
try: vparts[-1]=str(int(vparts[-1])+1)
except: vparts=['1','8','1']
NEWVER='.'.join(vparts)

new={'version':NEWVER,'tiers':TIERS,'volumeGuards':GUARDS,
     'volumes':OLD.get('volumes',{}),'curves':curves,'books':booksOut}
json.dump(new, open(OUT_JSON,'w'), separators=(',',':'))

# ---- report ----
old_keys=set(OLD['books']); new_keys=set(booksOut)
print('version',OLD.get('version'),'->',NEWVER)
print('books:',len(booksOut),'(current',len(old_keys),') | workbook-defined:',len(wb_books),'| from-dashboard-only preserved:',len(set(base_books)-set(wb_books)))
print('curves:',len(curves))
print('books LOST vs current:',len(old_keys-new_keys), list(old_keys-new_keys)[:8])
print('books ADDED vs current:',len(new_keys-old_keys), list(new_keys-old_keys)[:8])
print('guards:',len(GUARDS),'(current',len(OLD.get("volumeGuards",{})),') | gained:',sorted(set(GUARDS)-set(OLD.get('volumeGuards',{})))[:20])
print('volumes preserved:',len(new['volumes']))
print('CONFLICT dupes in workbook:',len([d for d in dupes if d[2]=='CONFLICT']))
print('bad tier cells:',len(badtier), badtier[:4])
